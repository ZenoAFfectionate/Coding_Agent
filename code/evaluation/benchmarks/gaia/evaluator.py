"""
GAIA Evaluator Module

Evaluates agent performance on the GAIA benchmark.
"""

from typing import Dict, Any, List, Optional, Union, Tuple
import time
import re
import json
import logging
import base64
from pathlib import Path
from code.evaluation.benchmarks.gaia.dataset import GAIADataset
from code.evaluation.benchmarks.gaia.metrics import GAIAMetrics

logger = logging.getLogger(__name__)

# Evaluation prompts directory (benchmark-specific prompts)
EVAL_PROMPTS_DIR = Path(__file__).resolve().parent.parent.parent / "prompts"


def _load_prompt(path: Path, fallback: str | None = None) -> str | None:
    """Load a prompt file, returning *fallback* if the file is missing or empty."""
    try:
        text = path.read_text(encoding="utf-8").strip()
        return text or fallback
    except FileNotFoundError:
        logger.warning("Prompt file not found: %s", path)
        return fallback


class GAIAEvaluator:
    """GAIA Evaluator

    Evaluates general AI assistant capabilities including:
    - Question understanding and reasoning
    - Multi-step problem solving
    - Tool use ability
    - Answer accuracy

    GAIA evaluation uses strict answer matching criteria:
    - Exact match: answer is identical
    - Partial match: answer contains correct information but differs in format

    Attributes:
        dataset: GAIA dataset
        metrics: Evaluation metrics calculator
        level: Difficulty level
        strict_mode: Whether to use strict matching mode
    """

    def __init__(
        self,
        dataset: Optional[GAIADataset] = None,
        level: Optional[int] = None,
        local_data_dir: Optional[str] = None,
        strict_mode: bool = True,
        llm=None,
    ):
        """Initialize the GAIA evaluator.

        Args:
            dataset: GAIA dataset; if None, one is created automatically
            level: Difficulty level (1-3)
            local_data_dir: Local data directory
            strict_mode: Whether to use strict matching mode
            llm: Optional HelloAgentsLLM instance for direct LLM mode
                 (bypasses the ReAct agent loop).
        """
        self.dataset = dataset or GAIADataset(
            level=level,
            local_data_dir=local_data_dir
        )
        self.metrics = GAIAMetrics()
        self.level = level
        self.strict_mode = strict_mode
        self.llm = llm

        # Load benchmark-specific prompts
        self.gaia_system_prompt = _load_prompt(EVAL_PROMPTS_DIR / "gaia_system.prompt")
        self.task_template = _load_prompt(EVAL_PROMPTS_DIR / "gaia_task.prompt")

    def evaluate(self, agent: Any = None, max_samples: Optional[int] = None) -> Dict[str, Any]:
        """Run evaluation.

        Args:
            agent: Agent to evaluate (used only when self.llm is None)
            max_samples: Maximum number of samples to evaluate; None means all

        Returns:
            Evaluation results dict with metrics
        """
        mode_label = "direct LLM" if self.llm else "agent"
        agent_name = getattr(self.llm, 'model', None) or getattr(agent, 'name', 'Unknown')

        print(f"\n🌟 Starting GAIA evaluation...")
        print(f"   Mode: {mode_label}")
        print(f"   Model/Agent: {agent_name}")
        print(f"   Difficulty level: {self.level or 'all'}")
        print(f"   Match mode: {'strict' if self.strict_mode else 'lenient'}")

        # Load dataset
        dataset = self.dataset.load()
        if not dataset:
            print("   ⚠️ Dataset is empty, skipping evaluation")
            return self._create_empty_results(agent)

        # Limit sample count
        if max_samples:
            dataset = dataset[:max_samples]

        print(f"   Sample count: {len(dataset)}")

        # Run evaluation
        results = []
        skipped_samples = 0
        level_stats = {1: {"total": 0, "correct": 0, "partial": 0},
                      2: {"total": 0, "correct": 0, "partial": 0},
                      3: {"total": 0, "correct": 0, "partial": 0}}

        for i, sample in enumerate(dataset):
            if i % 10 == 0:
                print(f"   Progress: {i+1}/{len(dataset)}")

            try:
                sample_result = self.evaluate_sample(agent, sample)
                results.append(sample_result)

                # Track skipped samples
                if sample_result.get("skipped"):
                    skipped_samples += 1
                    continue

                # Per-level statistics
                level = sample.get("level", 1)
                if level in level_stats:
                    level_stats[level]["total"] += 1
                    if sample_result["exact_match"]:
                        level_stats[level]["correct"] += 1
                    if sample_result["partial_match"]:
                        level_stats[level]["partial"] += 1

            except Exception as e:
                print(f"   ⚠️ Sample {i} evaluation failed: {e}")
                results.append({
                    "exact_match": False,
                    "partial_match": False,
                    "predicted": None,
                    "expected": sample.get("final_answer"),
                    "error": str(e),
                    "score": 0.0
                })

        # Calculate overall metrics (skipped samples excluded from totals)
        evaluated_results = [r for r in results if not r.get("skipped")]
        total_samples = len(evaluated_results)
        exact_matches = sum(1 for r in evaluated_results if r["exact_match"])
        partial_matches = sum(1 for r in evaluated_results if r["partial_match"])

        exact_match_rate = exact_matches / total_samples if total_samples > 0 else 0.0
        partial_match_rate = partial_matches / total_samples if total_samples > 0 else 0.0

        # Calculate per-level metrics
        level_metrics = {}
        for level, stats in level_stats.items():
            if stats["total"] > 0:
                level_metrics[f"Level_{level}"] = {
                    "total": stats["total"],
                    "exact_matches": stats["correct"],
                    "partial_matches": stats["partial"],
                    "exact_match_rate": stats["correct"] / stats["total"],
                    "partial_match_rate": stats["partial"] / stats["total"]
                }

        final_results = {
            "benchmark": "GAIA",
            "agent_name": agent_name,
            "strict_mode": self.strict_mode,
            "level_filter": self.level,
            "total_samples": total_samples,
            "skipped_samples": skipped_samples,
            "exact_matches": exact_matches,
            "partial_matches": partial_matches,
            "exact_match_rate": exact_match_rate,
            "partial_match_rate": partial_match_rate,
            "level_metrics": level_metrics,
            "detailed_results": results
        }

        print(f"✅ GAIA evaluation complete")
        print(f"   Evaluated samples: {total_samples} (skipped: {skipped_samples})")
        print(f"   Exact match rate: {exact_match_rate:.2%}")
        print(f"   Partial match rate: {partial_match_rate:.2%}")
        for level_name, metrics in level_metrics.items():
            print(f"   {level_name}: {metrics['exact_match_rate']:.2%} exact / {metrics['partial_match_rate']:.2%} partial")

        return final_results

    def evaluate_sample(self, agent: Any, sample: Dict[str, Any]) -> Dict[str, Any]:
        """Evaluate a single sample.

        Uses direct LLM invocation when self.llm is set, otherwise falls back
        to agent.run().  Loads file attachments and includes their content
        in the prompt (text) or as multimodal image blocks (LLM mode only).

        Args:
            agent: Agent to evaluate (ignored when self.llm is set)
            sample: Sample data

        Returns:
            Evaluation result for the single sample
        """
        try:
            # Prepare input
            question = sample.get("question", "")
            expected_answer = sample.get("final_answer", "")
            level = sample.get("level", 1)
            task_id = sample.get("task_id", "")

            # Load file attachment
            file_name = sample.get("file_name", "")
            content_type, file_content = self._load_file_content(file_name) if file_name else (None, None)

            # Skip unsupported file types
            if file_name and content_type is None and file_content is None:
                print(f"   ⏭️ Skipping sample {task_id}: unsupported file type ({Path(file_name).suffix})")
                return {
                    "task_id": task_id,
                    "level": level,
                    "exact_match": False,
                    "partial_match": False,
                    "score": 0.0,
                    "predicted": None,
                    "expected": expected_answer,
                    "skipped": True,
                    "skip_reason": f"Unsupported file type: {Path(file_name).suffix}",
                }

            # Build prompt with file text content (if applicable)
            file_text = file_content if content_type == "text" else None
            prompt = self._build_prompt(
                question, sample, file_text=file_text,
                image_attached=(content_type == "image"),
            )

            start_time = time.time()

            if self.llm:
                # Direct LLM mode — prepend system prompt if available
                messages = []
                if self.gaia_system_prompt:
                    messages.append({"role": "system", "content": self.gaia_system_prompt})
                if content_type == "image":
                    messages.extend(self._build_multimodal_messages(prompt, file_content, file_name))
                else:
                    messages.append({"role": "user", "content": prompt})
                response = self.llm.invoke(messages)
            else:
                # Agent mode — images not supported (agent.run takes a string)
                if content_type == "image":
                    print(f"   ⏭️ Skipping sample {task_id}: image files not supported in agent mode")
                    return {
                        "task_id": task_id,
                        "level": level,
                        "exact_match": False,
                        "partial_match": False,
                        "score": 0.0,
                        "predicted": None,
                        "expected": expected_answer,
                        "skipped": True,
                        "skip_reason": "Image files not supported in agent mode",
                    }
                response = agent.run(prompt)

            execution_time = time.time() - start_time

            # Extract answer
            predicted_answer = self._extract_answer(response)

            # Evaluate answer
            exact_match = self._check_exact_match(predicted_answer, expected_answer)
            partial_match = self._check_partial_match(predicted_answer, expected_answer)

            # Calculate score
            if exact_match:
                score = 1.0
            elif partial_match:
                score = 0.5
            else:
                score = 0.0

            return {
                "task_id": task_id,
                "level": level,
                "exact_match": exact_match,
                "partial_match": partial_match,
                "score": score,
                "predicted": predicted_answer,
                "expected": expected_answer,
                "response": response,
                "execution_time": execution_time
            }

        except Exception as e:
            return {
                "task_id": sample.get("task_id", ""),
                "level": sample.get("level", 1),
                "exact_match": False,
                "partial_match": False,
                "score": 0.0,
                "predicted": None,
                "expected": sample.get("final_answer", ""),
                "error": str(e)
            }

    def _create_empty_results(self, agent: Any) -> Dict[str, Any]:
        """Create empty evaluation results."""
        agent_name = getattr(self.llm, 'model', None) or getattr(agent, 'name', 'Unknown')
        return {
            "benchmark": "GAIA",
            "agent_name": agent_name,
            "strict_mode": self.strict_mode,
            "level_filter": self.level,
            "total_samples": 0,
            "skipped_samples": 0,
            "exact_matches": 0,
            "partial_matches": 0,
            "exact_match_rate": 0.0,
            "partial_match_rate": 0.0,
            "level_metrics": {},
            "detailed_results": []
        }

    def _build_prompt(
        self,
        question: str,
        sample: Dict[str, Any],
        file_text: Optional[str] = None,
        image_attached: bool = False,
    ) -> str:
        """Build the evaluation prompt.

        Uses the gaia_task.prompt template if available, otherwise falls back
        to the original inline format.

        Args:
            question: Question text
            sample: Sample data
            file_text: Pre-loaded file text content (for text-type attachments)
            image_attached: Whether an image is attached as a multimodal block
        """
        if self.task_template:
            file_section = ""
            if file_text and sample.get("file_name"):
                file_basename = Path(sample["file_name"]).name
                file_section = f"\n## Attached File: {file_basename}\n\n{file_text}\n"
            elif image_attached and sample.get("file_name"):
                file_section = f"\nAn image file ({Path(sample['file_name']).name}) is attached. Refer to it to answer the question.\n"
            elif sample.get("file_name") and not file_text:
                file_section = f"\nNote: This question may require reference to the file: {Path(sample['file_name']).name}\n"
            return self.task_template.format(
                question=question,
                file_section=file_section,
            )

        # Fallback: original inline prompt
        prompt = f"{question}"

        if file_text and sample.get("file_name"):
            file_basename = Path(sample["file_name"]).name
            prompt += f"\n\n## Attached File: {file_basename}\n\n{file_text}"
        elif image_attached and sample.get("file_name"):
            prompt += f"\n\nAn image file ({Path(sample['file_name']).name}) is attached. Refer to it to answer the question."
        elif sample.get("file_name") and not file_text:
            prompt += f"\n\nNote: This question may require reference to the file: {Path(sample['file_name']).name}"

        return prompt

    def _load_file_content(self, file_path: str) -> Tuple[Optional[str], Optional[str]]:
        """Read a file from disk and return its content for prompt inclusion.

        Returns:
            (content_type, content) where content_type is "text", "image", or
            None (unsupported).  For text, content is the string.  For image,
            content is a base64 data-URL.
        """
        if not file_path:
            return None, None

        path = Path(file_path)
        if not path.exists():
            print(f"   ⚠️ File not found: {file_path}")
            return None, None

        suffix = path.suffix.lower()

        # --- Text-readable formats ---
        text_extensions = {
            ".txt", ".md", ".py", ".csv", ".json", ".jsonl",
            ".xml", ".html", ".log", ".tsv", ".yaml", ".yml",
        }
        if suffix in text_extensions:
            try:
                content = path.read_text(encoding="utf-8")
                # Truncate very large files to avoid blowing up the context
                max_chars = 50_000
                if len(content) > max_chars:
                    content = content[:max_chars] + "\n\n... [truncated]"
                return "text", content
            except Exception as e:
                print(f"   ⚠️ Failed to read file ({path.name}): {e}")
                return None, None

        # --- Excel ---
        if suffix in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                text_parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text_parts.append(f"### Sheet: {sheet_name}\n")
                    for row in ws.iter_rows(values_only=True):
                        row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                        text_parts.append(row_str)
                wb.close()
                content = "\n".join(text_parts)
                max_chars = 50_000
                if len(content) > max_chars:
                    content = content[:max_chars] + "\n\n... [truncated]"
                return "text", content
            except ImportError:
                print("   ⚠️ openpyxl is not installed; cannot read Excel files")
                return None, None
            except Exception as e:
                print(f"   ⚠️ Failed to read Excel file ({path.name}): {e}")
                return None, None

        # --- PDF ---
        if suffix == ".pdf":
            try:
                import pypdf
                reader = pypdf.PdfReader(str(path))
                pages_text = []
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        pages_text.append(text)
                content = "\n\n".join(pages_text)
                max_chars = 50_000
                if len(content) > max_chars:
                    content = content[:max_chars] + "\n\n... [truncated]"
                return "text", content
            except ImportError:
                try:
                    import PyPDF2
                    reader = PyPDF2.PdfReader(str(path))
                    pages_text = []
                    for page in reader.pages:
                        text = page.extract_text()
                        if text:
                            pages_text.append(text)
                    content = "\n\n".join(pages_text)
                    max_chars = 50_000
                    if len(content) > max_chars:
                        content = content[:max_chars] + "\n\n... [truncated]"
                    return "text", content
                except ImportError:
                    print("   ⚠️ pypdf/PyPDF2 is not installed; cannot read PDF files")
                    return None, None
            except Exception as e:
                print(f"   ⚠️ Failed to read PDF file ({path.name}): {e}")
                return None, None

        # --- Image ---
        image_extensions = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
        if suffix in image_extensions:
            try:
                mime_map = {
                    ".png": "image/png",
                    ".jpg": "image/jpeg",
                    ".jpeg": "image/jpeg",
                    ".gif": "image/gif",
                    ".webp": "image/webp",
                    ".bmp": "image/bmp",
                }
                mime = mime_map.get(suffix, "image/png")
                raw = path.read_bytes()
                b64 = base64.b64encode(raw).decode("ascii")
                data_url = f"data:{mime};base64,{b64}"
                return "image", data_url
            except Exception as e:
                print(f"   ⚠️ Failed to read image ({path.name}): {e}")
                return None, None

        # --- Unsupported ---
        print(f"   ⚠️ Unsupported file type: {suffix}")
        return None, None

    def _build_multimodal_messages(self, prompt: str, image_data_url: str, file_name: str) -> List[Dict]:
        """Construct OpenAI-compatible multimodal messages with an image.

        Args:
            prompt: The text prompt.
            image_data_url: Base64 data URL for the image.
            file_name: Original file name (for logging context).

        Returns:
            Messages list suitable for HelloAgentsLLM.invoke().
        """
        return [{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": image_data_url}},
            ]
        }]

    def _extract_answer(self, response: str) -> str:
        """Extract the answer from the response (GAIA format).

        GAIA requires the answer format: FINAL ANSWER: [answer]
        """
        # First try to extract the official GAIA format answer
        final_answer_pattern = r'FINAL ANSWER:\s*(.+?)(?:\n|$)'
        match = re.search(final_answer_pattern, response, re.IGNORECASE | re.MULTILINE)
        if match:
            answer = match.group(1).strip()
            # Remove possible brackets
            answer = answer.strip('[]')
            return answer

        # Fallback: look for other answer markers
        answer_patterns = [
            r'Final answer[：:]\s*(.+)',
            r'Answer[：:]\s*(.+)',
        ]

        for pattern in answer_patterns:
            match = re.search(pattern, response, re.IGNORECASE)
            if match:
                return match.group(1).strip()

        # If no marker found, return the last non-empty line
        lines = response.strip().split('\n')
        for line in reversed(lines):
            line = line.strip()
            if line and not line.startswith('#'):
                return line

        return response.strip()

    def _check_exact_match(self, predicted: str, expected: str) -> bool:
        """Check for exact match."""
        if not predicted or not expected:
            return False

        # Normalize strings
        pred_normalized = self._normalize_answer(predicted)
        exp_normalized = self._normalize_answer(expected)

        return pred_normalized == exp_normalized

    def _check_partial_match(self, predicted: str, expected: str) -> bool:
        """Check for partial match."""
        if not predicted or not expected:
            return False

        # Normalize strings
        pred_normalized = self._normalize_answer(predicted)
        exp_normalized = self._normalize_answer(expected)

        # Check containment
        if exp_normalized in pred_normalized or pred_normalized in exp_normalized:
            return True

        # Check keyword overlap
        pred_words = set(pred_normalized.split())
        exp_words = set(exp_normalized.split())

        if not exp_words:
            return False

        # If over 70% of expected words appear in the prediction, consider it a partial match
        overlap = len(pred_words & exp_words)
        return overlap / len(exp_words) >= 0.7

    def _normalize_answer(self, answer: str) -> str:
        """Normalize answer string (GAIA official normalization rules).

        Based on the GAIA paper's normalization rules:
        1. Numbers: remove comma separators and unit symbols
        2. Strings: remove articles, lowercase, remove extra whitespace
        3. Lists: split by comma, normalize each element independently
        """
        if not answer:
            return ""

        answer = answer.strip()

        # Check if it's a comma-separated list
        if ',' in answer:
            # Split and normalize each element
            parts = [self._normalize_single_answer(p.strip()) for p in answer.split(',')]
            # Sort alphabetically (GAIA requirement)
            parts.sort()
            return ','.join(parts)
        else:
            return self._normalize_single_answer(answer)

    def _normalize_single_answer(self, answer: str) -> str:
        """Normalize a single answer (without commas)."""
        answer = answer.strip().lower()

        # Remove common articles
        articles = ['the', 'a', 'an']
        words = answer.split()
        if words and words[0] in articles:
            words = words[1:]
            answer = ' '.join(words)

        # Remove currency symbols and percent signs
        answer = answer.replace('$', '').replace('%', '').replace('€', '').replace('£', '')

        # Remove comma separators in numbers (e.g. 1,000 -> 1000)
        # but keep decimal points
        answer = re.sub(r'(\d),(\d)', r'\1\2', answer)

        # Remove extra whitespace
        answer = ' '.join(answer.split())

        # Remove trailing punctuation
        answer = answer.rstrip('.,;:!?')

        return answer

    def export_to_gaia_format(
        self,
        results: Dict[str, Any],
        output_path: Union[str, Path],
        include_reasoning: bool = True
    ) -> None:
        """Export results to GAIA official format.

        GAIA format requires:
        - JSONL format (one JSON object per line)
        - Each object contains: task_id, model_answer, reasoning_trace (optional)

        Args:
            results: Evaluation results
            output_path: Output file path
            include_reasoning: Whether to include reasoning traces
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        detailed_results = results.get("detailed_results", [])

        with open(output_path, 'w', encoding='utf-8') as f:
            for result in detailed_results:
                gaia_result = {
                    "task_id": result.get("task_id", ""),
                    "model_answer": result.get("predicted", "")
                }

                if include_reasoning:
                    gaia_result["reasoning_trace"] = result.get("response", "")

                f.write(json.dumps(gaia_result, ensure_ascii=False) + '\n')

        print(f"✅ GAIA format results exported")
        print(f"   Output file: {output_path}")
        print(f"   Samples: {len(detailed_results)}")
        print(f"   Include reasoning traces: {include_reasoning}")
